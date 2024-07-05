# chart.py

"""
This module is responsible for creating various types of charts in Excel workbooks using the openpyxl library.
It provides functionality to generate charts based on user-specified data ranges and chart types.
"""

# Import Libraries
import openpyxl
import logging
from openpyxl.chart import BarChart, LineChart, AreaChart, BubbleChart, RadarChart, PieChart, DoughnutChart, ScatterChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.chart.plotarea import DataTable


# Set up logging at the top of the file
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def create_chart(file_path, sheet_name, start_cell, end_cell, chart_type, chart_title, target_sheet_name=None):
    try:
        logging.info(f"Opening workbook: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        
        source_worksheet = workbook[sheet_name]
        target_worksheet = workbook[target_sheet_name] if target_sheet_name in workbook.sheetnames else source_worksheet

        start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
        end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)
        
        start_col_index = openpyxl.utils.column_index_from_string(start_col)
        end_col_index = openpyxl.utils.column_index_from_string(end_col)

        logging.info(f"Processing range: {start_cell} to {end_cell}")

        chart_classes = {
            "bar": BarChart, "line": LineChart, "area": AreaChart, "bubble": BubbleChart,
            "radar": RadarChart, "pie": PieChart, "doughnut": DoughnutChart, "scatter": ScatterChart
        }
        ChartClass = chart_classes.get(chart_type, BarChart)
        
        chart = ChartClass()
        chart.title = chart_title
        chart.style = 10

        data = Reference(source_worksheet, min_col=start_col_index + 1, min_row=start_row,
                         max_col=end_col_index, max_row=int(end_row))
        chart.add_data(data, titles_from_data=True)

        cats = Reference(source_worksheet, min_col=start_col_index, min_row=start_row + 1, max_row=int(end_row))
        chart.set_categories(cats)
        
        chart.legend = Legend()
        
        data_table = DataTable()
        chart.plot_area.dTable = data_table
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True
        
        chart.width = 16 + ((end_col_index - start_col_index + 1) * 0.8)
        chart.height = 8 + ((int(end_row) - start_row + 1) * 0.5)

        empty_row = 1
        while target_worksheet.cell(row=empty_row, column=1).value is not None:
            empty_row += 1

        target_worksheet.add_chart(chart, f'A{empty_row}')

        logging.info(f"Chart created successfully. Type: {chart_type}, Title: {chart_title}")
        
        logging.info("Saving workbook")
        workbook.save(file_path)
        logging.info("Workbook saved successfully")
        workbook.close()

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        raise